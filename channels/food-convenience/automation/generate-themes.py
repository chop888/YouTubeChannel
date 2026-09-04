"""毎週のテーマ案生成スクリプト（claude -p ヘッドレスモードを使用）

使い方:
    python generate-themes.py                 # 10件生成してtheme-stock.xlsxに追記
    python generate-themes.py --count 5        # 件数を変える
    python generate-themes.py --dry-run        # claudeを呼ばずダミーデータで動作確認
"""

import argparse
import json
import logging
import sys
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from claude_client import DEFAULT_CLAUDE_CMD, call_claude, extract_json_array, setup_logging

SCRIPT_DIR = Path(__file__).resolve().parent
CHANNEL_DIR = SCRIPT_DIR.parent
DEFAULT_XLSX = CHANNEL_DIR / "research" / "theme-stock.xlsx"
LOG_FILE = CHANNEL_DIR / "research" / "generate-themes.log"

HEADERS = ["日付", "型", "テーマ案", "選定理由", "ステータス"]
COLUMN_WIDTHS = [12, 18, 50, 55, 14]
STATUS_OPTIONS = ["未着手", "台本作成済み", "動画作成済み"]
STATUS_DROPDOWN_RANGE = "E2:E10000"

ALLOWED_TYPES = [
    "比較(2項目)",
    "比較(3項目)",
    "おすすめ〇選",
    "雑学系",
    "単品ディープダイブ",
    "定点観測",
]

BREAKING_TYPE = "速報ショート型"

logger = logging.getLogger("generate-themes")


def build_prompt(count: int, existing_themes: list[dict] | None = None) -> str:
    types_list = "\n".join(f"- {t}" for t in ALLOWED_TYPES)

    existing_block = ""
    if existing_themes:
        existing_lines = "\n".join(
            f"- ［{t['type']}］{t['theme']}" for t in existing_themes
        )
        existing_block = f"""
# 既存テーマ一覧（重複回避のため必ず確認すること）
theme-stock.xlsxには、すでに以下のテーマが登録されています。

{existing_lines}

- 上記と同一、またはほぼ同じテーマ（対象商品・切り口が同じもの）は提案しないこと
- 特に「単品ディープダイブ」では、上記一覧で既に扱った商品を再度取り上げないこと
- 「雑学系」「おすすめ〇選」では、切り口が既存テーマの傾向に偏らないよう注意すること
  （例：雑学系の既存テーマが「知られざる工夫」型ばかりに偏っていれば「知られざる不思議」型を
  多めにする、おすすめ〇選が「コスパ・お得」ばかりに偏っていれば「事件性・意外性」型を
  多めにするなど、全体のバランスを見て調整すること）
"""

    return f"""あなたは、YouTubeチャンネル「プチ得グルメラボ」（日本の食にまつわる雑学・発見チャンネル）の企画担当です。
次回の動画テーマ案を{count}件、考えてください。

# チャンネルの型（6種類）
以下のラベルを"type"の値としてそのまま使ってください。
{types_list}
{existing_block}
# テーマ選定の必須条件
- 価格や仕様など、事実確認・裏取りがしやすいテーマを優先すること
- 批判的・ネガティブな切り口（「まずい」「損する」等）は避け、ポジティブな切り口で成立するテーマにすること
- 極端にニッチで、内容が想像しにくいテーマは避けること
- 対象は「日本の食」全般とする：ご当地グルメ・郷土料理・伝統食・食べ物の歴史・昔の食べられ方・
  食べ物以外の意外な用途なども含む。コンビニ・スーパー・業務スーパーの商品や、マクドナルド・
  サイゼリヤ・牛丼チェーンなど外食チェーンのメニューも引き続き対象に含めてよいが、
  全体の中で最も登場頻度が低い切り口として扱うこと

# 型ごとの比重（均等配分ではなく、以下の目安に従うこと）
6種類の型は均等ではなく、明確に比重をつけること。{count}件の中の目安件数は以下（{count}=10の場合の例も併記）：
- 単品ディープダイブ：**最大の主力**。目安 全体の40%程度（10件なら4件）
- 雑学系：**主力**（ディープダイブに次ぐ比重）。目安 全体の30%程度（10件なら3件）
- 比較(2項目)＋比較(3項目)：たまに使う程度。目安 合計10%程度（10件なら合計1件）
- おすすめ〇選：たまに使う程度。目安 10%程度（10件なら1件）
- 定点観測：役割見直し中のため最小限。目安 10%程度（10件なら1件、無理に含めなくてもよい）

上記はあくまで目安であり、{count}が小さい場合や既存テーマの傾向によっては多少前後してよいが、
「単品ディープダイブ」が最も多く、次いで「雑学系」が多くなるようにし、この2つで全体の過半数を占めるようにすること。

# 型ごとの注意点
- 雑学系：動画1本につき関連する雑学ネタを3つ深掘りする構成のため、1つの狭い豆知識単体をテーマにしないこと。
  3つのネタを束ねられる程度の広さのテーマにすること。
  NG例：「おにぎりのフィルムが海苔をパリッと保つ仕組みの雑学」（ネタが1つしかない）
  OK例：「おにぎりにまつわる、知られざる工夫の雑学3選」（3つのネタを内包できる広さ）

  テーマの切り口は、大きく分けて以下の2パターンがある。概ね半々程度の比率になるよう意識し、
  毎回同じパターンに偏らないこと：
  ・「知られざる工夫」型：誰かが意図して設計した理由がある事象
    （例：おにぎりのフィルムの形状に込められた工夫）
  ・「知られざる不思議」型：理由がはっきりしない、謎めいた事象や不思議な現象を扱う
    （例：「なぜか〇〇の方が美味しく感じる」「〇〇の賞味期限の幅はなぜこんなに違うのか」
    のような、明確な設計意図というより「不思議だね」で終わるような話題も含める）

  切り口の例（あくまで発想のヒントであり、必ずこの通りにする必要はない）：
  ・ご当地グルメ・郷土料理にまつわる雑学（名前の由来、地域による味・呼び方の違いなど）
  ・食べ物の歴史（いつ・どこで生まれたか、時代によって食べられ方がどう変わったか）
  ・行事食の由来（おせち・七夕そうめん・土用の丑の日のうなぎ等、なぜその日に食べるのか）
  ・昔の保存食・知恵にまつわる雑学
  ・食品表示の意味（賞味期限と消費期限の違いなど）
  ・添加物や保存方法に関する豆知識
  ・栄養に関する意外な事実
  ・食品の製造工程の裏側
  ・食べ物以外の意外な用途があった話
  ・「なぜ〇〇なの？」「実は〇〇だった」という、日常で気づかない疑問を掘り下げる形式
    例：「なぜポテトチップスの袋はパンパンなの？」
        「なぜ牛乳は冷蔵庫のドアポケットに入れない方がいいと言われるのか」
        のような、身近な「そうだったの！？」を狙う切り口

  注意：添加物・保存方法など、健康や安全性に関わる話題を扱う場合は、不安を煽ったり、
  特定の商品・企業を批判したりする表現は避け、「知ると面白い豆知識」として中立的・
  ポジティブなトーンを保つこと。根拠が曖昧な情報は扱わず、公的機関や企業公式の情報源を優先する。

- おすすめ〇選：テーマの切り口は、大きく分けて以下の2パターンがある。両方をバランスよく
  含め、毎回「コスパ・お得」ばかりに偏らないこと：
  ・「おすすめ」型：前向きな推薦（コスパが良い、リピートしたくなる、など）
  ・「事件性・意外性」型：販売終了になった伝説の商品、予想外に人気が出すぎた商品、
    開発秘話の中の意外なエピソードなど、話題性・意外性のあるエピソードを扱う
    （例：「二度と食べられない伝説の〇選」「まさかここまで人気が出るとは思わなかった〇選」）

  「ご当地グルメ〇選」「郷土料理〇選」のような、地域の食にまつわる切り口も歓迎する。
  コンビニ・スーパー商品中心の回は、この型の中でも最も低頻度に留めること。

  注意：「事件性・意外性」型を扱う場合も、特定の商品・企業を批判・中傷する表現は避け、
  あくまで「面白い事実・エピソード」としてポジティブなトーンで扱うこと。

- 単品ディープダイブ：1つの食べ物・料理を多角的に掘り下げる型のため、切り口の幅がテーマの
  面白さを左右する。ご当地グルメ・郷土料理・歴史的な食べ物を主な対象としつつ、マクドナルド・
  サイゼリヤ・牛丼チェーンなど外食チェーンの定番メニューや、コンビニ・スーパー・業務スーパーの
  商品も対象に含めてよい（コンビニ・スーパー商品は最も低頻度に留める）。
  同じ商品・料理を繰り返し取り上げないよう、既存テーマ一覧（上記）を必ず確認すること。

  この型の中でも、対象の種類ごとに比重をつけること：
  ・**「料理・食べ物そのもの」が主役のテーマ**（例：長崎ちゃんぽん、肉まん、おでん等、
    特定の1社に紐づかない料理・食品ジャンルとしての深掘り）：目安7〜8割
  ・**「特定の企業・商品」が主役のテーマ**：目安2〜3割。このカテゴリで想定しているのは、
    **外食チェーンの看板・定番メニュー**（マクドナルド・サイゼリヤ・牛丼チェーン等）と、
    551蓬莱・白い恋人のような**地域性の強い老舗ブランド**が中心
  単品ディープダイブが複数件になる場合は、「料理・食べ物そのもの」が主役のテーマを
  優先的に多めにすること。

  注意：日清「チキンラーメン」、亀田製菓「柿の種」のような、**コンビニ・スーパーで売っている
  大手メーカーの加工食品・お菓子は「特定の企業・商品」カテゴリとしては扱わない**。取り上げる
  場合は、その商品名を主役にするのではなく、「料理・食べ物そのもの」側の切り口に寄せて
  再構成すること（例：「チキンラーメン」ではなく「即席麺の歴史」、「柿の種」ではなく
  「米菓の歴史」のように、食品ジャンルとしての深掘りにする）。

  切り口の例（あくまで発想のヒントであり、必ずこの通りにする必要はない）：
  ・誕生の経緯・開発秘話、発祥の地
  ・時代によって食べられ方がどう変わってきたか
  ・形状や仕様に込められた理由
  ・製造工程の裏側
  ・栄養成分の特徴
  ・おすすめの食べ方・アレンジ
  ・意外な歴史やエピソード

- 比較(2項目)・比較(3項目)：地域による同じ料理の違い（例：関東と関西での味付けの違い）などの
  比較も積極的に含めてよい。コンビニ・スーパーの商品比較は、この型の中でも最も低頻度に留めること。

- 定点観測：役割を見直し中の型。コンビニ・スーパー商品の値上げ・値下げ追跡よりも、
  ご当地名産品・特産品の価格変動を追跡する使い方を優先的に検討すること。
  適切なテーマが見当たらない週は、無理にこの型を含めなくてよい。

上記の切り口はあくまで発想のヒントであり、このプロンプトで示した条件（ポジティブなトーン、
事実確認のしやすさ、固有名詞の裏取りなど）と矛盾しない範囲で、新しい切り口も自由に考えてよい。
なお、あなたはファイルシステムへのアクセスやツールの使用を一切行わないこと。追加の確認や
ファイル参照が必要だと感じても、許可を求める文面は書かず、このプロンプトに書かれている
情報だけをもとに判断すること。

# 出力形式（厳守）
説明文・前置き・後書きは一切書かず、以下の形式のJSON配列のみを出力してください。
マークダウンのコードブロック（```）も使わないでください。

[
  {{"type": "型のラベル", "theme": "テーマ案（1文）", "reason": "選定理由（1〜2文）"}}
]
"""


def build_news_prompt(max_items: int) -> str:
    today = date.today()
    week_ago = today - timedelta(days=7)
    return f"""あなたは、YouTubeチャンネル「プチ得グルメラボ」（日本の食にまつわる雑学・発見チャンネル）の企画担当です。
Web検索を使って、食品業界全般に関する直近1週間以内のニュースがないか確認してください。

# 検索対象期間
本日は{today.isoformat()}です。{week_ago.isoformat()}以降に発表・報道された情報のみを対象にしてください。

# 検索の観点（固定キーワードではなく、幅広く検索すること）
チャンネルのコンセプト（コンビニ・スーパー・業務スーパーの商品・食のお得情報）に関連する、
価格変動や新商品の動きを幅広く拾ってください。検索例（これに限定しない）：
「コンビニ 値下げ」「スーパー 値下げ」「食品 値上げ」「米価格」
「コンビニ 新商品」「業務スーパー 新商品」「コンビニ おにぎり」など

# 採用条件
- 実際にWeb検索で見つかった、実在する情報のみを使うこと（推測・創作は禁止）
- 各項目には、検索で見つけた実際の情報源のURLを必ず含めること
- 複数のニュースが見つかった場合は、視聴者の実際の買い物に直接関係が深いもの
  （コンビニ・スーパーでの値下げ・値上げ・新商品など）を優先し、最大{max_items}件までに絞ること
- 期間外（{week_ago.isoformat()}より前）の情報や、業界と無関係な一般ニュースは含めないこと
- 該当する情報が1件も見つからない場合は、空配列を返すこと（無理に何か含めないこと）

# 出力形式（厳守）
説明文・前置き・後書きは一切書かず、以下の形式のJSON配列のみを出力してください（該当なしの場合は []）。
マークダウンのコードブロック（```）も使わないでください。

[
  {{"theme": "動画の題材となる一言（例：〇〇が値下げ、〇〇が新発売、のような形）", "reason": "ニュースの概要と情報源URLを含む1〜2文"}}
]
"""


def parse_breaking_news(result_text: str, max_items: int) -> list[dict]:
    cleaned = extract_json_array(result_text)
    try:
        items = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"速報ニュースのJSONを解析できませんでした。内容:\n{cleaned[:2000]}"
        ) from exc

    if not isinstance(items, list):
        raise RuntimeError(f"速報ニュースがリスト形式ではありませんでした: {cleaned[:500]}")

    valid = []
    for i, item in enumerate(items):
        if not isinstance(item, dict) or not all(k in item for k in ("theme", "reason")):
            logger.warning("速報ニュースの形式不正のためスキップ（%d件目）: %s", i, item)
            continue
        if "http" not in item["reason"]:
            logger.warning("速報ニュースにURLが含まれていない可能性があります（%d件目）: %s", i, item)
        valid.append({"type": BREAKING_TYPE, "theme": item["theme"], "reason": item["reason"]})

    return valid[:max_items]


def check_breaking_news(timeout: int, claude_cmd: str, max_items: int) -> list[dict]:
    prompt = build_news_prompt(max_items)
    result_text = call_claude(prompt, timeout, claude_cmd, extra_args='--allowedTools "WebSearch"')
    return parse_breaking_news(result_text, max_items)


def parse_themes(result_text: str) -> list[dict]:
    cleaned = extract_json_array(result_text)
    try:
        items = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"テーマ案のJSONを解析できませんでした。内容:\n{cleaned[:2000]}"
        ) from exc

    if not isinstance(items, list):
        raise RuntimeError(f"テーマ案がリスト形式ではありませんでした: {cleaned[:500]}")

    valid = []
    for i, item in enumerate(items):
        if not isinstance(item, dict) or not all(k in item for k in ("type", "theme", "reason")):
            logger.warning("形式不正のためスキップ（%d件目）: %s", i, item)
            continue
        if item["type"] not in ALLOWED_TYPES:
            logger.warning("未知の型ラベルですが採用します（%d件目）: %s", i, item["type"])
        valid.append(item)

    if not valid:
        raise RuntimeError("有効なテーマ案が1件も得られませんでした。")

    return valid


def load_existing_themes(xlsx_path: Path) -> list[dict]:
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    themes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3 or not row[2]:
            continue
        themes.append({"type": str(row[1] or ""), "theme": str(row[2])})
    return themes


def dummy_themes(count: int) -> list[dict]:
    return [
        {
            "type": ALLOWED_TYPES[i % len(ALLOWED_TYPES)],
            "theme": f"[ダミー] テーマ案サンプル{i + 1}",
            "reason": "[ダミー] --dry-run実行時のテストデータです。",
        }
        for i in range(count)
    ]


def ensure_status_dropdown(ws) -> None:
    formula = '"' + ",".join(STATUS_OPTIONS) + '"'
    for dv in ws.data_validations.dataValidation:
        if dv.formula1 == formula:
            return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "リストから選択してください"
    dv.errorTitle = "入力エラー"
    dv.add(STATUS_DROPDOWN_RANGE)
    ws.add_data_validation(dv)


def append_to_xlsx(path: Path, themes: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "theme-stock"
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i, width in enumerate(COLUMN_WIDTHS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        ws.freeze_panes = "A2"

    ensure_status_dropdown(ws)

    today = date.today()
    for item in themes:
        row = [today, item["type"], item["theme"], item["reason"], STATUS_OPTIONS[0]]
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).number_format = "yyyy-mm-dd"

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="週次テーマ案生成スクリプト")
    parser.add_argument("--count", type=int, default=10, help="生成するテーマ案の件数（既定: 10）")
    parser.add_argument("--output", type=Path, default=DEFAULT_XLSX, help="出力するxlsxのパス")
    parser.add_argument("--timeout", type=int, default=180, help="claude呼び出しのタイムアウト秒数")
    parser.add_argument("--news-timeout", type=int, default=240, help="速報ニュース検索のタイムアウト秒数")
    parser.add_argument("--max-breaking", type=int, default=3, help="速報ショート型として採用する最大件数")
    parser.add_argument("--skip-news", action="store_true", help="速報ニュース検索をスキップする")
    parser.add_argument("--dry-run", action="store_true", help="claudeを呼ばずダミーデータで動作確認する")
    parser.add_argument(
        "--claude-cmd",
        default=DEFAULT_CLAUDE_CMD,
        help=f"claude実行ファイルのパス（既定: {DEFAULT_CLAUDE_CMD}）",
    )
    args = parser.parse_args()

    global logger
    logger = setup_logging(LOG_FILE, "generate-themes")
    logger.info("=== テーマ案生成開始（count=%d, dry_run=%s） ===", args.count, args.dry_run)

    try:
        breaking_items = []
        if not args.skip_news:
            try:
                breaking_items = check_breaking_news(args.news_timeout, args.claude_cmd, args.max_breaking)
                breaking_items = breaking_items[: args.count]
                if breaking_items:
                    logger.info("速報ニュースを%d件検出しました。", len(breaking_items))
                else:
                    logger.info("該当する速報ニュースは見つかりませんでした。")
            except Exception as exc:
                logger.warning("速報ニュース検索に失敗したため、通常のテーマ生成のみ行います: %s", exc)
                breaking_items = []

        remaining = max(args.count - len(breaking_items), 0)

        try:
            existing_themes = load_existing_themes(args.output)
        except Exception as exc:
            logger.warning("既存テーマの読み込みに失敗したため、重複チェックなしで生成します: %s", exc)
            existing_themes = []

        if args.dry_run:
            # ニュース検索は上で実際に実行済み。ここから先（通常テーマ生成）だけダミーにする。
            evergreen_items = dummy_themes(remaining)
        else:
            evergreen_items = []
            if remaining > 0:
                prompt = build_prompt(remaining, existing_themes)
                result_text = call_claude(prompt, args.timeout, args.claude_cmd)
                evergreen_items = parse_themes(result_text)

        themes = breaking_items + evergreen_items

        append_to_xlsx(args.output, themes)
        logger.info("%d件のテーマ案を %s に追記しました。", len(themes), args.output)
        logger.info("=== テーマ案生成成功 ===")
        return 0
    except Exception as exc:
        logger.error("テーマ案生成に失敗しました: %s", exc)
        logger.info("=== テーマ案生成失敗 ===")
        return 1


if __name__ == "__main__":
    sys.exit(main())
