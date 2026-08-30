"""theme-stock.xlsx を使って台本を自動生成するスクリプト（claude -p ヘッドレスモードを使用）

使い方:
    python generate-script.py                    # ステータスが未着手の中から最古の1件を選んで生成
    python generate-script.py --theme "テーマ名"   # テーマを直接指定して生成
    python generate-script.py --dry-run           # claudeを呼ばずダミー台本で動作確認
"""

import argparse
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from claude_client import DEFAULT_CLAUDE_CMD, call_claude, setup_logging, strip_code_fence

SCRIPT_DIR = Path(__file__).resolve().parent
CHANNEL_DIR = SCRIPT_DIR.parent
SCRIPTS_DIR = CHANNEL_DIR / "scripts"
DEFAULT_XLSX = CHANNEL_DIR / "research" / "theme-stock.xlsx"
LOG_FILE = CHANNEL_DIR / "research" / "generate-script.log"
TEMPLATES_DIR = CHANNEL_DIR / "docks" / "script-templates"
COMMON_OPENING_FILE = TEMPLATES_DIR / "00_common_opening_template.md"
CHARACTERS_FILE = CHANNEL_DIR / "docks" / "02-characters.md"
STYLE_SAMPLE_FILE = SCRIPTS_DIR / "comparison-maratang.md"

STATUS_NOT_STARTED = "未着手"
STATUS_SCRIPTED = "台本作成済み"

TYPE_TEMPLATE_FILES = {
    "比較(2項目)": "comparison-2items.md",
    "比較(3項目)": "comparison-3items.md",
    "おすすめ〇選": "ranking.md",
    "雑学系": "trivia.md",
    "単品ディープダイブ": "deepdive.md",
    "定点観測": "pricewatch.md",
    "速報ショート型": "breaking-short.md",
}

TYPE_SLUGS = {
    "比較(2項目)": "comparison2",
    "比較(3項目)": "comparison3",
    "おすすめ〇選": "ranking",
    "雑学系": "trivia",
    "単品ディープダイブ": "deepdive",
    "定点観測": "pricewatch",
    "速報ショート型": "breaking",
}

TYPE_MARKER_RE = re.compile(r"<!--\s*type:\s*(.+?)\s*-->\s*\n*")

logger = None


def load_text(path: Path) -> str:
    if not path.exists():
        raise RuntimeError(f"必要なファイルが見つかりません: {path}")
    return path.read_text(encoding="utf-8")


def read_theme_rows(xlsx_path: Path) -> list[dict]:
    if not xlsx_path.exists():
        raise RuntimeError(
            f"{xlsx_path} が見つかりません。先にgenerate-themes.pyでテーマ案を生成してください。"
        )
    wb = load_workbook(xlsx_path)
    ws = wb.active
    rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[2] is None:
            continue
        rows.append(
            {
                "row_idx": idx,
                "date": row[0],
                "type": row[1],
                "theme": row[2],
                "reason": row[3],
                "status": row[4],
            }
        )
    return rows


def _date_sort_key(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    return datetime.min


def select_target(rows: list[dict], theme_arg: str | None) -> tuple[dict | None, str]:
    """戻り値: (theme-stock.xlsxの該当行 または None, テーマ文字列)"""
    if theme_arg:
        for r in rows:
            if r["theme"] == theme_arg:
                return r, theme_arg
        return None, theme_arg

    candidates = [r for r in rows if r["status"] == STATUS_NOT_STARTED]
    if not candidates:
        raise RuntimeError(
            f"ステータスが「{STATUS_NOT_STARTED}」のテーマがtheme-stock.xlsxにありません。"
        )
    candidates.sort(key=lambda r: _date_sort_key(r["date"]))
    chosen = candidates[0]
    return chosen, chosen["theme"]


CHAR_TARGET_ROW_RE = re.compile(
    r"^\|\s*(?P<section>[^|]+?)\s*\|\s*(?P<time>[^|]+?)\s*\|\s*(?P<chars>[^|]*字[^|]*?)\s*\|\s*[^|]*\|\s*$",
    re.MULTILINE,
)

def extract_char_targets(template_text: str) -> str:
    lines = []
    for m in CHAR_TARGET_ROW_RE.finditer(template_text):
        section = m.group("section").strip().strip("*").strip()
        chars = m.group("chars").strip().strip("*").strip()
        lines.append(f"- {section}：{chars}")
    return "\n".join(lines)


def build_known_type_prompt(
    theme: str, type_label: str, template_text: str, opening_text: str, characters_text: str, style_sample: str
) -> str:
    char_targets = extract_char_targets(template_text)
    return f"""あなたは、YouTubeチャンネル「プチ得ラボ」（コンビニ・スーパー・業務スーパーの商品比較チャンネル）の台本作家です。
以下の情報だけをもとに、今回のテーマの台本を1本、Markdown形式で書いてください。

# 今回のテーマ
{theme}

# 今回使用する型
{type_label}

# 各パートの目標文字数（必ず守ること）
えりか・ここなの実測読み上げ速度は平均約346文字/分。過去の生成では台本が短くなりすぎたため、
以下の目標文字数（句読点・記号を含む、話者名は除いたセリフ本文の文字数）を各パートで必ず満たすこと。

{char_targets}

## 目標文字数を満たすための具体的な書き方
文字数だけを意識しても足りなくなりやすいため、以下の目安で掛け合いのボリューム自体を増やすこと。
- 1パートにつき、えりかとここなの掛け合いを7〜9往復（セリフ合計14〜18個程度）ほど書くこと
- えりかのセリフは1回あたり2〜3文（目安60〜120字）にし、単発の短い一言で終わらせないこと
- ここなのセリフも1〜2文（目安30〜80字）にし、「そうなんだ」のような相槌だけで終わらせず、
  自分の感想・具体的なエピソード・追加の疑問などを添えること
- 「価格→特徴→リアクション」で終わらせず、背景・理由・具体例・比較の掘り下げを1往復ずつ追加すること

# 台本テンプレート（この構成・時間配分・セリフの型に厳密に従うこと）
{template_text}

# 固定オープニング（導入セクションの一番先頭に、このセリフをそのまま必ず挿入すること）
{opening_text}

# キャラクター設定（口調・一人称を必ず反映すること）
{characters_text}

# 書式のお手本（過去の台本サンプル。Markdownの書式・体裁だけを踏襲すること。内容は今回のテーマに差し替える）
{style_sample}

# 執筆時の注意
- あなたはファイルの書き込み・保存を一切行わないこと。ツールは使わず、テキストで直接回答すること。
  この回答本文がそのままプログラムによって台本ファイルとして保存されるため、
  「保存の許可」や「保存先の確認」などのやり取りは一切不要（そのような文面も書かないこと）
- 冒頭に示した「各パートの目標文字数」を必ず守ること。相槌だけの短いセリフを連続させて終わらせず、
  説明・リアクション・具体例を厚めに書いて分量を確保すること
- 実際の価格や数値データが不明な場合は「◯◯円」のようにダミー表記にし、末尾の「制作メモ」欄に差し替え箇所を明記すること
- 企業の商品シリーズ名・キャンペーン名・製法名などの固有名詞を使う場合は、実在するか確信が持てなくても構わないので、
  その部分の直後に「※要確認」を必ず付けること（例：「おにぎり屋シリーズ※要確認」）。これは公開前の裏取りチェック漏れを防ぐためのマーカーであり、
  固有名詞そのものの使用を妨げるものではない
- キャラクターは「えりか」「ここな」の2人のみ
- セリフは「**えりか**：」「**ここな**：」のように、話者名を太字（**）で囲んだMarkdown形式にすること（書式のお手本と同じ書き方）
- 固定オープニングの直後に、テーマに応じた導入（お題提示・つかみ）を続けること
- 説明文・前置き・後書きは一切書かず、台本のMarkdown本文のみを出力すること（```で全体を囲まないこと）
"""


def build_unknown_type_prompt(
    theme: str, all_templates: dict[str, str], opening_text: str, characters_text: str, style_sample: str
) -> str:
    templates_block = "\n\n".join(f"## 型：{label}\n{text}" for label, text in all_templates.items())
    char_targets_block = "\n\n".join(
        f"【{label}】\n{extract_char_targets(text)}" for label, text in all_templates.items()
    )
    return f"""あなたは、YouTubeチャンネル「プチ得ラボ」（コンビニ・スーパー・業務スーパーの商品比較チャンネル）の台本作家です。
以下のテーマは、まだ台本の型が決まっていません。6種類の型テンプレートの中から最も適したものを1つ選び、
その型の構成に厳密に従って台本を1本、Markdown形式で書いてください。

# 今回のテーマ
{theme}

# 型ごとの目標文字数（必ず守ること）
えりか・ここなの実測読み上げ速度は平均約346文字/分。過去の生成では台本が短くなりすぎたため、
選んだ型の目標文字数（句読点・記号を含む、話者名は除いたセリフ本文の文字数）を各パートで必ず満たすこと。

{char_targets_block}

## 目標文字数を満たすための具体的な書き方
文字数だけを意識しても足りなくなりやすいため、以下の目安で掛け合いのボリューム自体を増やすこと。
- 1パートにつき、えりかとここなの掛け合いを7〜9往復（セリフ合計14〜18個程度）ほど書くこと
- えりかのセリフは1回あたり2〜3文（目安60〜120字）にし、単発の短い一言で終わらせないこと
- ここなのセリフも1〜2文（目安30〜80字）にし、「そうなんだ」のような相槌だけで終わらせず、
  自分の感想・具体的なエピソード・追加の疑問などを添えること
- 「価格→特徴→リアクション」で終わらせず、背景・理由・具体例・比較の掘り下げを1往復ずつ追加すること

# 型テンプレート一覧（この中から1つだけ選ぶこと）
{templates_block}

# 固定オープニング（導入セクションの一番先頭に、このセリフをそのまま必ず挿入すること）
{opening_text}

# キャラクター設定（口調・一人称を必ず反映すること）
{characters_text}

# 書式のお手本（過去の台本サンプル。Markdownの書式・体裁だけを踏襲すること。内容は今回のテーマに差し替える）
{style_sample}

# 出力形式（厳守）
出力の1行目には、選んだ型を必ず次の形式で書くこと（ラベルは「型テンプレート一覧」の「型：」の右側の文字列と完全に一致させること）。
<!-- type: 選んだ型のラベル -->

2行目は空行にし、3行目以降に台本のMarkdown本文を書くこと。

# 執筆時の注意
- あなたはファイルの書き込み・保存を一切行わないこと。ツールは使わず、テキストで直接回答すること。
  この回答本文がそのままプログラムによって台本ファイルとして保存されるため、
  「保存の許可」や「保存先の確認」などのやり取りは一切不要（そのような文面も書かないこと）
- 冒頭に示した「各パートの目標文字数」を必ず守ること。相槌だけの短いセリフを連続させて終わらせず、
  説明・リアクション・具体例を厚めに書いて分量を確保すること
- 実際の価格や数値データが不明な場合は「◯◯円」のようにダミー表記にし、末尾の「制作メモ」欄に差し替え箇所を明記すること
- 企業の商品シリーズ名・キャンペーン名・製法名などの固有名詞を使う場合は、実在するか確信が持てなくても構わないので、
  その部分の直後に「※要確認」を必ず付けること（例：「おにぎり屋シリーズ※要確認」）。これは公開前の裏取りチェック漏れを防ぐためのマーカーであり、
  固有名詞そのものの使用を妨げるものではない
- キャラクターは「えりか」「ここな」の2人のみ
- セリフは「**えりか**：」「**ここな**：」のように、話者名を太字（**）で囲んだMarkdown形式にすること（書式のお手本と同じ書き方）
- 固定オープニングの直後に、テーマに応じた導入（お題提示・つかみ）を続けること
- 1行目のマーカー以外に、説明文・前置き・後書きは一切書かず、台本のMarkdown本文のみを出力すること（```で全体を囲まないこと）
"""


def extract_type_marker(text: str) -> tuple[str | None, str]:
    match = TYPE_MARKER_RE.match(text.strip())
    if not match:
        return None, text.strip()
    label = match.group(1).strip()
    remainder = text.strip()[match.end() :].strip()
    return label, remainder


def slugify_theme(theme: str, max_len: int = 24) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]', "", theme)
    cleaned = cleaned.replace(" ", "").replace("　", "")
    return cleaned[:max_len] or "theme"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem, suffix = path.stem, path.suffix
    i = 2
    while True:
        candidate = path.with_name(f"{stem}-{i}{suffix}")
        if not candidate.exists():
            return candidate
        i += 1


def dummy_script(theme: str, type_label: str) -> str:
    return f"""# 台本サンプル：{theme}

- 型：{type_label}
- [ダミー] --dry-run実行時のテストデータです。
"""


def update_status(xlsx_path: Path, row_idx: int) -> None:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    status_cell = ws.cell(row=row_idx, column=5)
    if status_cell.value != STATUS_NOT_STARTED:
        return
    status_cell.value = STATUS_SCRIPTED
    wb.save(xlsx_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="theme-stock.xlsxを使った台本自動生成スクリプト")
    parser.add_argument("--theme", type=str, default=None, help="生成するテーマ名を直接指定する")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="theme-stock.xlsxのパス")
    parser.add_argument("--output-dir", type=Path, default=SCRIPTS_DIR, help="台本の保存先ディレクトリ")
    parser.add_argument("--timeout", type=int, default=300, help="claude呼び出しのタイムアウト秒数")
    parser.add_argument("--dry-run", action="store_true", help="claudeを呼ばずダミー台本で動作確認する")
    parser.add_argument(
        "--claude-cmd",
        default=DEFAULT_CLAUDE_CMD,
        help=f"claude実行ファイルのパス（既定: {DEFAULT_CLAUDE_CMD}）",
    )
    args = parser.parse_args()

    global logger
    logger = setup_logging(LOG_FILE, "generate-script")
    logger.info("=== 台本生成開始（theme=%s, dry_run=%s） ===", args.theme, args.dry_run)

    try:
        rows = read_theme_rows(args.xlsx)
        row, theme = select_target(rows, args.theme)

        opening_text = load_text(COMMON_OPENING_FILE)
        characters_text = load_text(CHARACTERS_FILE)
        style_sample = load_text(STYLE_SAMPLE_FILE)

        if row is not None:
            type_label = row["type"]
            if type_label not in TYPE_TEMPLATE_FILES:
                raise RuntimeError(f"theme-stock.xlsxの型ラベルが未知です: {type_label}")
            logger.info(
                "テーマ「%s」（型: %s）をtheme-stock.xlsxの%d行目から選択しました。",
                theme,
                type_label,
                row["row_idx"],
            )
        else:
            type_label = None
            logger.info(
                "テーマ「%s」はtheme-stock.xlsxに見つからないため、型をclaudeに判断させます。", theme
            )

        if args.dry_run:
            chosen_type = type_label or "比較(2項目)"
            script_text = dummy_script(theme, chosen_type)
        elif type_label is not None:
            template_text = load_text(TEMPLATES_DIR / TYPE_TEMPLATE_FILES[type_label])
            prompt = build_known_type_prompt(
                theme, type_label, template_text, opening_text, characters_text, style_sample
            )
            result_text = call_claude(prompt, args.timeout, args.claude_cmd)
            script_text = strip_code_fence(result_text)
            chosen_type = type_label
        else:
            all_templates = {
                label: load_text(TEMPLATES_DIR / filename) for label, filename in TYPE_TEMPLATE_FILES.items()
            }
            prompt = build_unknown_type_prompt(theme, all_templates, opening_text, characters_text, style_sample)
            result_text = call_claude(prompt, args.timeout, args.claude_cmd)
            result_text = strip_code_fence(result_text)
            marker_type, script_text = extract_type_marker(result_text)
            if marker_type in TYPE_TEMPLATE_FILES:
                chosen_type = marker_type
                logger.info("claudeが型を「%s」と判断しました。", chosen_type)
            else:
                logger.warning("claudeが返した型ラベルを認識できませんでした: %s", marker_type)
                chosen_type = "theme"

        stripped_head = script_text.lstrip()
        if not (stripped_head.startswith("#") or stripped_head.startswith("**")):
            raise RuntimeError(
                "claudeの出力が台本の形式（見出し、または太字の話者名で開始）になっていません。"
                f"保存確認の文面などが混ざっている可能性があります。先頭200文字:\n{script_text[:200]}"
            )

        slug = TYPE_SLUGS.get(chosen_type, "theme")
        filename = f"{slug}-{slugify_theme(theme)}.md"
        args.output_dir.mkdir(parents=True, exist_ok=True)
        output_path = unique_path(args.output_dir / filename)
        output_path.write_text(script_text.strip() + "\n", encoding="utf-8")
        logger.info("台本を %s に保存しました。", output_path)

        if row is not None and not args.dry_run:
            update_status(args.xlsx, row["row_idx"])
            logger.info(
                "theme-stock.xlsxの%d行目のステータスを「%s」に更新しました。", row["row_idx"], STATUS_SCRIPTED
            )

        logger.info("=== 台本生成成功 ===")
        return 0
    except Exception as exc:
        logger.error("台本生成に失敗しました: %s", exc)
        logger.info("=== 台本生成失敗 ===")
        return 1


if __name__ == "__main__":
    sys.exit(main())
