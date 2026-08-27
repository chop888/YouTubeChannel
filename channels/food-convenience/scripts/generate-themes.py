"""毎週のテーマ案生成スクリプト（claude -p ヘッドレスモードを使用）

使い方:
    python generate-themes.py                 # 10件生成してtheme-stock.xlsxに追記
    python generate-themes.py --count 5        # 件数を変える
    python generate-themes.py --dry-run        # claudeを呼ばずダミーデータで動作確認
"""

import argparse
import json
import logging
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from claude_client import DEFAULT_CLAUDE_CMD, call_claude, setup_logging, strip_code_fence

SCRIPT_DIR = Path(__file__).resolve().parent
CHANNEL_DIR = SCRIPT_DIR.parent
DEFAULT_XLSX = CHANNEL_DIR / "research" / "theme-stock.xlsx"
LOG_FILE = CHANNEL_DIR / "research" / "generate-themes.log"

HEADERS = ["日付", "型", "テーマ案", "選定理由", "ステータス"]
COLUMN_WIDTHS = [12, 18, 50, 55, 14]
STATUS_OPTIONS = ["未着手", "台本作成済み", "動画作成済み"]
STATUS_DROPDOWN_RANGE = "E2:E10000"

ALLOWED_TYPES = [
    "比較(2項目)",
    "比較(3項目)",
    "おすすめ〇選",
    "雑学系",
    "単品ディープダイブ",
    "定点観測",
]

logger = logging.getLogger("generate-themes")


def build_prompt(count: int) -> str:
    types_list = "\n".join(f"- {t}" for t in ALLOWED_TYPES)
    return f"""あなたは、YouTubeチャンネル「プチ得ラボ」（コンビニ・スーパー・業務スーパーの商品比較チャンネル）の企画担当です。
次回の動画テーマ案を{count}件、考えてください。

# チャンネルの型（6種類）
以下のラベルを"type"の値としてそのまま使ってください。
{types_list}

# テーマ選定の必須条件
- 価格や仕様など、事実確認・裏取りがしやすいテーマを優先すること
- 批判的・ネガティブな切り口（「まずい」「損する」等）は避け、ポジティブな切り口で成立するテーマにすること
- 上記6種類の型を、{count}件の中でできるだけ均等に含めること
- 極端にニッチで、商品イメージが想像しにくいテーマは避けること
- 対象はコンビニ（セブン・ローソン・ファミマ等）、スーパー、業務スーパーの商品・サービス

# 型ごとの注意点
- 雑学系：動画1本につき関連する雑学ネタを3つ深掘りする構成のため、1つの狭い豆知識単体をテーマにしないこと。
  3つのネタを束ねられる程度の広さのテーマにすること。
  NG例：「コンビニおにぎりのフィルムが海苔をパリッと保つ仕組みの雑学」（ネタが1つしかない）
  OK例：「コンビニおにぎりにまつわる、知られざる工夫の雑学3選」（3つのネタを内包できる広さ）

# 出力形式（厳守）
説明文・前置き・後書きは一切書かず、以下の形式のJSON配列のみを出力してください。
マークダウンのコードブロック（```）も使わないでください。

[
  {{"type": "型のラベル", "theme": "テーマ案（1文）", "reason": "選定理由（1〜2文）"}}
]
"""


def parse_themes(result_text: str) -> list[dict]:
    cleaned = strip_code_fence(result_text)
    try:
        items = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        raise RuntimeError(
            f"テーマ案のJSONを解析できませんでした。内容:\n{cleaned[:2000]}"
        ) from exc

    if not isinstance(items, list):
        raise RuntimeError(f"テーマ案がリスト形式ではありませんでした: {cleaned[:500]}")

    valid = []
    for i, item in enumerate(items):
        if not isinstance(item, dict) or not all(k in item for k in ("type", "theme", "reason")):
            logger.warning("形式不正のためスキップ（%d件目）: %s", i, item)
            continue
        if item["type"] not in ALLOWED_TYPES:
            logger.warning("未知の型ラベルですが採用します（%d件目）: %s", i, item["type"])
        valid.append(item)

    if not valid:
        raise RuntimeError("有効なテーマ案が1件も得られませんでした。")

    return valid


def dummy_themes(count: int) -> list[dict]:
    return [
        {
            "type": ALLOWED_TYPES[i % len(ALLOWED_TYPES)],
            "theme": f"[ダミー] テーマ案サンプル{i + 1}",
            "reason": "[ダミー] --dry-run実行時のテストデータです。",
        }
        for i in range(count)
    ]


def ensure_status_dropdown(ws) -> None:
    formula = '"' + ",".join(STATUS_OPTIONS) + '"'
    for dv in ws.data_validations.dataValidation:
        if dv.formula1 == formula:
            return
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "リストから選択してください"
    dv.errorTitle = "入力エラー"
    dv.add(STATUS_DROPDOWN_RANGE)
    ws.add_data_validation(dv)


def append_to_xlsx(path: Path, themes: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "theme-stock"
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for i, width in enumerate(COLUMN_WIDTHS, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
        ws.freeze_panes = "A2"

    ensure_status_dropdown(ws)

    today = date.today()
    for item in themes:
        row = [today, item["type"], item["theme"], item["reason"], STATUS_OPTIONS[0]]
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).number_format = "yyyy-mm-dd"

    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="週次テーマ案生成スクリプト")
    parser.add_argument("--count", type=int, default=10, help="生成するテーマ案の件数（既定: 10）")
    parser.add_argument("--output", type=Path, default=DEFAULT_XLSX, help="出力するxlsxのパス")
    parser.add_argument("--timeout", type=int, default=180, help="claude呼び出しのタイムアウト秒数")
    parser.add_argument("--dry-run", action="store_true", help="claudeを呼ばずダミーデータで動作確認する")
    parser.add_argument(
        "--claude-cmd",
        default=DEFAULT_CLAUDE_CMD,
        help=f"claude実行ファイルのパス（既定: {DEFAULT_CLAUDE_CMD}）",
    )
    args = parser.parse_args()

    global logger
    logger = setup_logging(LOG_FILE, "generate-themes")
    logger.info("=== テーマ案生成開始（count=%d, dry_run=%s） ===", args.count, args.dry_run)

    try:
        if args.dry_run:
            themes = dummy_themes(args.count)
        else:
            prompt = build_prompt(args.count)
            result_text = call_claude(prompt, args.timeout, args.claude_cmd)
            themes = parse_themes(result_text)

        append_to_xlsx(args.output, themes)
        logger.info("%d件のテーマ案を %s に追記しました。", len(themes), args.output)
        logger.info("=== テーマ案生成成功 ===")
        return 0
    except Exception as exc:
        logger.error("テーマ案生成に失敗しました: %s", exc)
        logger.info("=== テーマ案生成失敗 ===")
        return 1


if __name__ == "__main__":
    sys.exit(main())
